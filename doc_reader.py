"""
doc_reader.py — Leitor e analisador de documentos multi-formato
Suporta: PDF, TXT, DOCX, CSV, XLSX
"""

import csv
import io
import logging
import os

logger = logging.getLogger("doc_reader")

# Imports opcionais — falham graciosamente
try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False
    logger.warning("pdfplumber não instalado — PDF desabilitado")

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    logger.warning("python-docx não instalado — DOCX desabilitado")

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    logger.warning("openpyxl não instalado — XLSX desabilitado")


SUPPORTED_EXTENSIONS = {
    ".pdf": "PDF",
    ".txt": "Texto",
    ".docx": "Word",
    ".csv": "CSV",
    ".xlsx": "Excel",
    ".md": "Markdown",
    ".json": "JSON",
    ".py": "Python",
    ".js": "JavaScript",
    ".log": "Log",
}


def extract_text(file_path: str) -> str:
    """
    Extrai texto de qualquer formato suportado.
    
    Returns:
        Texto extraído (limitado a 10000 chars pra não estourar LLM context)
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext not in SUPPORTED_EXTENSIONS:
        return f"Formato '{ext}' não suportado. Formatos aceitos: {', '.join(SUPPORTED_EXTENSIONS.keys())}"

    try:
        if ext == ".pdf":
            text = _read_pdf(file_path)
        elif ext == ".docx":
            text = _read_docx(file_path)
        elif ext == ".csv":
            text = _read_csv(file_path)
        elif ext == ".xlsx":
            text = _read_xlsx(file_path)
        else:
            # TXT, MD, JSON, PY, JS, LOG — tudo é texto
            text = _read_text(file_path)

        # Limitar tamanho
        if len(text) > 10000:
            text = text[:10000] + f"\n\n[... documento truncado, {len(text)} chars total]"

        logger.info(f"📄 Extraído {len(text)} chars de {os.path.basename(file_path)}")
        return text

    except Exception as e:
        logger.error(f"Erro lendo {file_path}: {e}")
        return f"Erro ao ler arquivo: {str(e)[:200]}"


def _read_pdf(path: str) -> str:
    """Extrai texto de PDF via pdfplumber."""
    if not HAS_PDF:
        return "pdfplumber não está instalado. Rode: pip install pdfplumber"

    text_parts = []
    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and text.strip():
                text_parts.append(f"--- Página {page_num + 1} ---\n{text}")

    return "\n\n".join(text_parts) if text_parts else "PDF sem texto extraível (pode ser escaneado/imagem)."


def _read_docx(path: str) -> str:
    """Extrai texto de DOCX via python-docx."""
    if not HAS_DOCX:
        return "python-docx não está instalado. Rode: pip install python-docx"

    doc = DocxDocument(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "Documento DOCX vazio."


def _read_csv(path: str) -> str:
    """Lê CSV e formata como tabela legível."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return "CSV vazio."

    # Header + primeiras 50 linhas
    header = rows[0]
    data_rows = rows[1:51]

    lines = [" | ".join(header)]
    lines.append("-" * len(lines[0]))
    for row in data_rows:
        lines.append(" | ".join(row))

    if len(rows) > 51:
        lines.append(f"\n[... {len(rows) - 51} linhas adicionais omitidas]")

    return "\n".join(lines)


def _read_xlsx(path: str) -> str:
    """Lê XLSX e formata como tabela legível."""
    if not HAS_XLSX:
        return "openpyxl não está instalado. Rode: pip install openpyxl"

    wb = load_workbook(path, read_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=51, values_only=True))

        if not rows:
            continue

        lines = [f"### Planilha: {sheet_name}"]
        for i, row in enumerate(rows):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
            if i == 0:
                lines.append("-" * len(lines[-1]))

        total_rows = ws.max_row or 0
        if total_rows > 51:
            lines.append(f"\n[... {total_rows - 51} linhas adicionais omitidas]")

        parts.append("\n".join(lines))

    wb.close()
    return "\n\n".join(parts) if parts else "Planilha vazia."


def _read_text(path: str) -> str:
    """Lê qualquer arquivo de texto."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


async def analyze_document(file_path: str, router, question: str = None) -> str:
    """
    Analisa um documento: extrai texto, pede ao LLM para resumir/analisar.
    
    Args:
        file_path: Caminho do arquivo
        router: LLMRouter
        question: Pergunta específica sobre o documento (opcional)
    """
    text = extract_text(file_path)
    filename = os.path.basename(file_path)

    if text.startswith("Erro") or text.startswith("Formato"):
        return text

    prompt_base = (
        f"Analise este documento ({filename}) e forneça:\n"
        "1. **Resumo** (2-3 parágrafos)\n"
        "2. **Pontos-chave** (lista)\n"
        "3. **Dados importantes** (se houver números/estatísticas)\n"
        "4. **Observações** (algo notável ou que merece atenção)\n\n"
        "Seja factual e cite trechos específicos quando relevante."
    )

    if question:
        prompt_base = (
            f"O Criador enviou o documento '{filename}' e perguntou: {question}\n\n"
            "Responda a pergunta com base no conteúdo do documento. "
            "Se a resposta não estiver no documento, diga isso."
        )

    result = await router.generate([
        {"role": "system", "content": prompt_base},
        {"role": "user", "content": f"Conteúdo do documento:\n\n{text}"},
    ], temperature=0.1)

    return result if isinstance(result, str) else "Erro na análise do documento."
